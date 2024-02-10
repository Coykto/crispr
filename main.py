from functools import cached_property

from openpyxl import load_workbook, Workbook


class ReadExcel:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(self.file_path)
        self.ws = self.wb.active

    @cached_property
    def headers(self):
        return [cell for cell in next(self.ws.iter_rows(values_only=True))]

    @cached_property
    def data(self):
        result = []
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            row_data = {self.headers[i]: value for i, value in enumerate(row)}
            result.append(row_data)
        return result


class WriteExcel:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = Workbook()
        self.ws = self.wb.active

    def write_rows(self, reader: ReadExcel):
        self.ws.append(reader.headers)
        for row_data in reader.data:
            row = [row_data.get(header) for header in reader.headers]
            self.ws.append(row)
        self.wb.save(self.file_path)


if __name__ == '__main__':
    reader = ReadExcel('input.xlsx')
    writer = WriteExcel('output.xlsx')
    writer.write_rows(reader)